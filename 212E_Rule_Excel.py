import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
import os


# List of countries subject to teh 212E rule
COUNTRIES_212E = {
    "Belize", "Benin", "Burkina Faso", "Burma", "Cabo Verde", "Cambodia",
    "Cameroon", "Democratic Republic of the Congo", "Djibouti", "Ecuador",
    "El Salvador", "Eritrea", "Ethiopia", "Fiji", "Gambia, The", "Ghana",
    "Guatemala", "Haiti", "Honduras", "Jamaica", "Kenya", "Kosovo", "Lebanon",
    "Liberia", "Malawi", "Mali", "Mauritania", "Mozambique", "Nepal",
    "Nicaragua", "Niger", "Nigeria", "Palestinian Authority (West Bank and Gaza)",
    "Philippines", "Rwanda", "Senegal", "Tajikistan", "Tanzania", "Timor-Leste",
    "Togo", "Tonga", "Venezuela", "Yemen", "Zambia"
}

def normalize_country(country):
    if not country:
        return ""
    return str(country).strip().lower()

# Opens little window to allow us t sleect a file
def select_file():
    # open file dialog to sleect excel file
    root= tk.Tk()
    root.withdraw() 

    file_path = filedialog.askopenfilename(
        title="Select Excel File",
        filetypes = [("Excel files", "*.xlsx *.xls")]
    )

    root.destroy
    return file_path

def process_excel(file_path):

    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        ws["I1"]  = "Updated subject to 212E"

        normalized_countries = {normalize_country(c) for c in COUNTRIES_212E}

        updates_count = 0

        #process rows
        for row in range(2, ws.max_row+1):
            citizenship = ws[f'E{row}'].value
            subject_212e = ws[f'B{row}'].value

            norm_citizenship = normalize_country(citizenship)

            # So basically if the subject is listed as YES they are subject AND their country is not
            # Listed under the skill list by country then we add the new field
            # Increase update count
            if subject_212e and str(subject_212e).strip().upper() == "YES":
                if norm_citizenship not in normalized_countries:
                    ws[f'I{row}'] = "No Longer subject"
                    updates_count += 1

        # save with a new file so i dont lose the old one/overwrite
        base, ext = os.path.splitext(file_path)
        new_file_path = f"{base}_updated{ext}"
        wb.save(new_file_path)

        return new_file_path, updates_count
    
    except Exception as e:
        raise Exception(f"Error processing file: {str(e)}")
    
def main():
    """Main function"""
    print("=== 212E Rule Excel Updater ===\n")
    print("This program will:")
    print("1. Let you select an Excel file")
    print("2. Add a new column 'Updated subject to 212E' in column H")
    print("3. Mark students as 'No longer subject' if their country is not on the updated list\n")

    input("Press Enter to select your Excel file...")
    
    # Select file
    file_path = select_file()
    
    if not file_path:
        print("No file selected. Exiting.")
        return
    
    print(f"\nSelected file: {file_path}")
    print("Processing...")
    
    try:
        new_file_path, updates_count = process_excel(file_path)
        print(f"\n✓ Success!")
        print(f"  - Updated {updates_count} student(s)")
        print(f"  - New file saved as: {new_file_path}")
        
        # Show success dialog
        root = tk.Tk()
        root.withdraw()
        messagebox.showinfo(
            "Success",
            f"File updated successfully!\n\n"
            f"Students updated: {updates_count}\n"
            f"New file: {os.path.basename(new_file_path)}"
        )
        root.destroy()
        
    except Exception as e:
        print(f"\n✗ Error: {str(e)}")
        root = tk.Tk()
        root.withdraw()
        messagebox.showerror("Error", str(e))
        root.destroy()

if __name__ == "__main__":
    main()