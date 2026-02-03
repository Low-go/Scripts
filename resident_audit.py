import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
from openpyxl import Workbook, load_workbook
import os

# Select which automated task you wish to perform
def select_task():
    result = {"value": None}  

    def choose(val):
        result["value"] = val
        root.destroy()  

    def on_close():
        root.destroy()

    root = tk.Tk()
    root.title("Select Job Automation")
    root.geometry("500x200")

    root.protocol("WM_DELETE_WINDOW", on_close)

    button1 = tk.Button(root, text="Combine housing with Main", command=lambda: choose(0))
    button1.pack(padx=20, pady=10)

    button2 = tk.Button(root, text="Format Housing Sheet", command=lambda: choose(1))
    button2.pack(padx=20, pady=10)

    button3 = tk.Button(root, text="Split Complete & Incomplete", command=lambda: choose(2))
    button3.pack(padx=20, pady=10)

    root.mainloop()  
    return result["value"]

# Select File window
def select_file():
    root = tk.Tk()
    root.withdraw()

    file_path = filedialog.askopenfilename(
        title= "Select Excel Files",
        filetypes=[("Excel Files", "*.xlsx *.xls")]
    )

    root.destroy()
    return file_path

# Convert column letter to index (A=0, B=1, etc.)
def col_letter_to_index(letter):
    letter = letter.upper().strip()
    result = 0
    for char in letter:
        result = result * 26 + (ord(char) - ord('A') + 1)
    return result - 1

# Get column mappings from user, this whole thing is to make it so this aint hardcoded
def get_column_mapping():
    result = {"mapping": None}
    
    def submit():
        try:
            # grab all the values and convert to indices
            mapping = {
                'table1': {
                    'student_id': col_letter_to_index(t1_id_entry.get()),
                    'last_name': col_letter_to_index(t1_last_entry.get()),
                    'first_name': col_letter_to_index(t1_first_entry.get()),
                    'address1': col_letter_to_index(t1_addr1_entry.get()),
                    'address2': col_letter_to_index(t1_addr2_entry.get()),
                    'city': col_letter_to_index(t1_city_entry.get()),
                    'postal': col_letter_to_index(t1_postal_entry.get())
                },
                'table2': {
                    'student_id': col_letter_to_index(t2_id_entry.get()),
                    'address1': col_letter_to_index(t2_addr1_entry.get()),
                    'address2': col_letter_to_index(t2_addr2_entry.get()),
                    'city': col_letter_to_index(t2_city_entry.get()),
                    'postal': col_letter_to_index(t2_postal_entry.get()),
                    'state': col_letter_to_index(t2_state_entry.get())
                }
            }
            result["mapping"] = mapping
            root.destroy()
        except Exception as e:
            messagebox.showerror("Error", f"Invalid column input: {e}")
    
    def cancel():
        root.destroy()
    
    root = tk.Tk()
    root.title("Specify Column Mappings")
    root.geometry("400x550")
    
    # Table 1 section
    tk.Label(root, text="Table 1 (Main) Columns:", font=("Arial", 11, "bold")).pack(pady=(10,5))
    
    frame1 = tk.Frame(root)
    frame1.pack(pady=5)
    
    tk.Label(frame1, text="Student ID:", width=12, anchor="w").grid(row=0, column=0, padx=5, pady=3)
    t1_id_entry = tk.Entry(frame1, width=10)
    t1_id_entry.grid(row=0, column=1, padx=5, pady=3)
    t1_id_entry.insert(0, "A")
    
    tk.Label(frame1, text="Last Name:", width=12, anchor="w").grid(row=1, column=0, padx=5, pady=3)
    t1_last_entry = tk.Entry(frame1, width=10)
    t1_last_entry.grid(row=1, column=1, padx=5, pady=3)
    t1_last_entry.insert(0, "B")
    
    tk.Label(frame1, text="First Name:", width=12, anchor="w").grid(row=2, column=0, padx=5, pady=3)
    t1_first_entry = tk.Entry(frame1, width=10)
    t1_first_entry.grid(row=2, column=1, padx=5, pady=3)
    t1_first_entry.insert(0, "C")
    
    tk.Label(frame1, text="Address 1:", width=12, anchor="w").grid(row=3, column=0, padx=5, pady=3)
    t1_addr1_entry = tk.Entry(frame1, width=10)
    t1_addr1_entry.grid(row=3, column=1, padx=5, pady=3)
    t1_addr1_entry.insert(0, "T")
    
    tk.Label(frame1, text="Address 2:", width=12, anchor="w").grid(row=4, column=0, padx=5, pady=3)
    t1_addr2_entry = tk.Entry(frame1, width=10)
    t1_addr2_entry.grid(row=4, column=1, padx=5, pady=3)
    t1_addr2_entry.insert(0, "U")
    
    tk.Label(frame1, text="City:", width=12, anchor="w").grid(row=5, column=0, padx=5, pady=3)
    t1_city_entry = tk.Entry(frame1, width=10)
    t1_city_entry.grid(row=5, column=1, padx=5, pady=3)
    t1_city_entry.insert(0, "G")
    
    tk.Label(frame1, text="Postal:", width=12, anchor="w").grid(row=6, column=0, padx=5, pady=3)
    t1_postal_entry = tk.Entry(frame1, width=10)
    t1_postal_entry.grid(row=6, column=1, padx=5, pady=3)
    t1_postal_entry.insert(0, "H")
    
    # Table 2 section
    tk.Label(root, text="Table 2 (Housing Audit) Columns:", font=("Arial", 11, "bold")).pack(pady=(15,5))
    
    frame2 = tk.Frame(root)
    frame2.pack(pady=5)
    
    tk.Label(frame2, text="Student ID:", width=12, anchor="w").grid(row=0, column=0, padx=5, pady=3)
    t2_id_entry = tk.Entry(frame2, width=10)
    t2_id_entry.grid(row=0, column=1, padx=5, pady=3)
    t2_id_entry.insert(0, "B")
    
    tk.Label(frame2, text="Address 1:", width=12, anchor="w").grid(row=1, column=0, padx=5, pady=3)
    t2_addr1_entry = tk.Entry(frame2, width=10)
    t2_addr1_entry.grid(row=1, column=1, padx=5, pady=3)
    t2_addr1_entry.insert(0, "C")
    
    tk.Label(frame2, text="Address 2:", width=12, anchor="w").grid(row=2, column=0, padx=5, pady=3)
    t2_addr2_entry = tk.Entry(frame2, width=10)
    t2_addr2_entry.grid(row=2, column=1, padx=5, pady=3)
    t2_addr2_entry.insert(0, "D")
    
    tk.Label(frame2, text="City:", width=12, anchor="w").grid(row=3, column=0, padx=5, pady=3)
    t2_city_entry = tk.Entry(frame2, width=10)
    t2_city_entry.grid(row=3, column=1, padx=5, pady=3)
    t2_city_entry.insert(0, "E")
    
    tk.Label(frame2, text="Postal:", width=12, anchor="w").grid(row=4, column=0, padx=5, pady=3)
    t2_postal_entry = tk.Entry(frame2, width=10)
    t2_postal_entry.grid(row=4, column=1, padx=5, pady=3)
    t2_postal_entry.insert(0, "F")

    tk.Label(frame2, text="State:", width=12, anchor="w").grid(row=5, column=0, padx=5, pady=3)
    t2_state_entry = tk.Entry(frame2, width=10)
    t2_state_entry.grid(row=5, column=1, padx=5, pady=3)
    t2_state_entry.insert(0, "G")
    
    # Buttons
    button_frame = tk.Frame(root)
    button_frame.pack(pady=20)
    
    submit_btn = tk.Button(button_frame, text="Submit", command=submit, width=10)
    submit_btn.pack(side="left", padx=10)
    
    cancel_btn = tk.Button(button_frame, text="Cancel", command=cancel, width=10)
    cancel_btn.pack(side="right", padx=10)
    
    root.mainloop()
    return result["mapping"]


def audit(file_path):

    try:
        wb = openpyxl.load_workbook(file_path)
        new_wb = Workbook()
        new_ws = new_wb.active
        ws = wb.worksheets[0]

        new_row_counter = 1

        # This section just grabs the student Names and student ids of the Resident excel
        # And moves them to column A and B of the new one
        for row in ws.iter_rows(min_row=2):

            col_I = row[8] 
            col_B = row[1]
            col_C = row[2]
            col_N = row[13]
            col_Q = row[16]

            if col_I.value == "CURRENT" or col_I.value == "ASSIGNED" or col_I.value == "INCOMING":

                if "Hale" in str(col_N.value):
                    address1 = "55-220 Kulanui St"
                    #remove leading zero if it exists
                    q_value = str(col_Q.value)
                    if q_value.startswith("0"):
                        q_value = q_value[1:]
                    address2 = f"H{q_value}"
                elif "TVA" in str(col_N.value):
                    address1 = "55-550 Naniloa Loop" 
                    address2 = f"TVA {col_Q.value}"
                else: 
                    address1 = col_N.value # fallbback
                    address2 = col_Q.value


                # new_row = col_I.row
                new_ws.cell(row=new_row_counter, column=1).value = col_B.value
                new_ws.cell(row=new_row_counter, column=2).value = col_C.value
                new_ws.cell(row=new_row_counter, column=3).value = address1
                new_ws.cell(row=new_row_counter, column=4).value = address2
                new_ws.cell(row=new_row_counter, column=5).value = "Laie"
                new_ws.cell(row=new_row_counter, column=6).value = "96762"
                new_ws.cell(row=new_row_counter, column=7).value = "Hawaii"

                new_row_counter +=1

        base, ext = os.path.splitext(file_path)
        new_file_path = f"{base}_updated{ext}"
        new_wb.save(new_file_path)

        
        return new_file_path

    except Exception as e:
            print(f"Error processing {file_path}: {e}")
            return None

def combine_audit_to_main(file1, file2, col_map):
    
    print(f"DEBUG: Starting combine with col_map = {col_map}")

    try:
        # Load up and open workesheet of both files
        # Make a new workbook where we will save this info in
        wb1 = openpyxl.load_workbook(file1)
        wb2 = openpyxl.load_workbook(file2)
        ws1 = wb1.worksheets[1]
        ws2 = wb2.worksheets[0]

        new_wb = Workbook()
        new_ws = new_wb.active
        new_ws.append(['Student Id', 'Student Last Name', 'Student First Name', 'Address 1', 'Address 2', 'City', 'Postal'])

        # Build dictionary/Hashmap for Table 2
        # Key = student id, Value ["address 1", "address 2", "City", "Postal"]
        table2_data = {}
        for row in ws2.iter_rows(min_row=2, values_only=True):
            student_id = row[col_map['table2']['student_id']]
            address1 = row[col_map['table2']['address1']]
            address2 = row[col_map['table2']['address2']]
            city2 = row[col_map['table2']['city']]
            postal2 = row[col_map['table2']['postal']]

            if student_id:
                table2_data[student_id] = {
                    'address1': address1,
                    'address2': address2,
                    'city2': city2,
                    'postal2': postal2
                }
        
        # Loop through table 1 make decisions
        for row in ws1.iter_rows(min_row=2, values_only=True):
            student_id = row[col_map['table1']['student_id']]
            student_last_name = row[col_map['table1']['last_name']]
            student_first_name = row[col_map['table1']['first_name']]

            # NOTE These will probably be changed in the future
            table1_address1 = row[col_map['table1']['address1']]
            table1_address2 = row[col_map['table1']['address2']]
            table1_city = row[col_map['table1']['city']]
            table1_postal = row[col_map['table1']['postal']]

            if student_id in table2_data:
                table2_info = table2_data[student_id]

                # Check if both addresses filled. NOTE might change this later
                if table2_info['address1'] and table2_info['address2']:
                    #Use table 2 address

                    new_ws.append([student_id, student_last_name, student_first_name, table2_info['address1'], table2_info['address2'], table2_info['city2'], table2_info['postal2']])
                else:
                    new_ws.append([student_id, student_last_name, student_first_name, 
                    table1_address1, 
                    table1_address2,
                    table1_city,
                    table1_postal
                    ])
            else:
                    new_ws.append([student_id, student_last_name, student_first_name,
                    table1_address1, 
                    table1_address2,
                    table1_city,
                    table1_postal
                    ])

        base, ext = os.path.splitext(file2)
        new_file_path = f"{base}_finalized_outcome{ext}"
        new_wb.save(new_file_path)

        
        return new_file_path


    except Exception as e:
        print(f"Error processing {file1} or {file2}: {e}")
        return None



def split_complete_incomplete(file1, file2, col_map):
    
    print(f"DEBUG: Starting split with col_map = {col_map}")

    try:
        # Load up and open workesheet of both files
        # Make two new workbooks - one for complete, one for incomplete
        wb1 = openpyxl.load_workbook(file1)
        wb2 = openpyxl.load_workbook(file2)
        ws1 = wb1.worksheets[1]
        ws2 = wb2.worksheets[0]

        # Workbook for complete entries (Table 2 data)
        complete_wb = Workbook()
        complete_ws = complete_wb.active
        complete_ws.append(['Student Id', 'Student Last Name', 'Student First Name', 'Address 1', 'Address 2', 'City', 'Postal', 'State'])

        # Workbook for incomplete entries (Table 1 fallback)
        incomplete_wb = Workbook()
        incomplete_ws = incomplete_wb.active
        incomplete_ws.append(['Student Id', 'Student Last Name', 'Student First Name', 'Address 1', 'Address 2', 'City', 'Postal'])

        # Build dictionary/Hashmap for Table 2
        # Key = student id, Value ["address 1", "address 2", "City", "Postal"]
        table2_data = {}
        for row in ws2.iter_rows(min_row=2, values_only=True):
            student_id = row[col_map['table2']['student_id']]
            address1 = row[col_map['table2']['address1']]
            address2 = row[col_map['table2']['address2']]
            city2 = row[col_map['table2']['city']]
            postal2 = row[col_map['table2']['postal']]
            state2 = row[col_map['table2']['state']]

            if student_id:
                table2_data[student_id] = {
                    'address1': address1,
                    'address2': address2,
                    'city2': city2,
                    'postal2': postal2,
                    'state2': state2 
                }
        
        # Loop through table 1 make decisions
        for row in ws1.iter_rows(min_row=2, values_only=True):
            student_id = row[col_map['table1']['student_id']]
            student_last_name = row[col_map['table1']['last_name']]
            student_first_name = row[col_map['table1']['first_name']]

            # NOTE These will probably be changed in the future
            table1_address1 = row[col_map['table1']['address1']]
            table1_address2 = row[col_map['table1']['address2']]
            table1_city = row[col_map['table1']['city']]
            table1_postal = row[col_map['table1']['postal']]

            if student_id in table2_data:
                table2_info = table2_data[student_id]

                # Check if both addresses filled. NOTE might change this later
                if table2_info['address1'] and table2_info['address2']:
                    # Use table 2 address - goes to COMPLETE sheet
                    complete_ws.append([student_id, student_last_name, student_first_name, 
                                       table2_info['address1'], table2_info['address2'], 
                                       table2_info['city2'], table2_info['postal2'], table2_info['state2']])
                else:
                    # Table 2 incomplete - goes to INCOMPLETE sheet
                    incomplete_ws.append([student_id, student_last_name, student_first_name, 
                                         table1_address1, table1_address2,
                                         table1_city, table1_postal])
            else:
                # Not in Table 2 - goes to INCOMPLETE sheet
                incomplete_ws.append([student_id, student_last_name, student_first_name,
                                     table1_address1, table1_address2,
                                     table1_city, table1_postal])

        base, ext = os.path.splitext(file2)
        complete_file_path = f"{base}_complete{ext}"
        incomplete_file_path = f"{base}_incomplete{ext}"
        
        complete_wb.save(complete_file_path)
        incomplete_wb.save(incomplete_file_path)

        return complete_file_path, incomplete_file_path

    except Exception as e:
        print(f"Error processing {file1} or {file2}: {e}")
        return None, None
    
    
def main():

    choice = select_task()

    if choice is None:
        print("No task selected")
        return

    # We format the housing excel sheet
    if choice == 1:
    
        file_path = select_file()

        if not file_path:
            print("No file sleected")
            return
        
        print(f"\nSelected file: {file_path}")
        print("Processing...")

        try:

            new_file_path = audit(file_path)
            print(f"\n Success!")
            print(f" - New file saved as : {new_file_path}")

            #show success dialog
            root = tk.Tk()
            root.withdraw()
            messagebox.showinfo(
                "Sucess",
                f"File updated successfully!\n\nNew file: {os.path.basename(new_file_path)}"
            )
            root.destroy()
        
        except Exception as e:
            print(f"\n✗ Error: {str(e)}")
            root = tk.Tk()
            root.withdraw()
            messagebox.showerror("Error", str(e))
            root.destroy()
    
    # We split into complete and incomplete tables
    elif choice == 2:
        print(f"\n Please select the main table first")
        print(f"\n ..................") 
        file_path1 = select_file()

        if not file_path1:
            print("No file sleected")
            return
        
        print(f"\n Now select the corrected Housing Audit Workbook")
        print(f"\n ..................")
        file_path2 = select_file()

        if not file_path2:
            print("No file sleected")
            return
        
        # Get column mappings from user
        col_map = get_column_mapping()

        if not col_map:
            print("Operation cancelled")
            return
        
        try:
            # Creates two files - complete and incomplete
            complete_path, incomplete_path = split_complete_incomplete(file_path1, file_path2, col_map)
            print(f"\n Success!")
            print(f" - Complete entries saved as: {complete_path}")
            print(f" - Incomplete entries saved as: {incomplete_path}")

            #show success dialog
            root = tk.Tk()
            root.withdraw()
            messagebox.showinfo(
                "Success",
                f"Files created successfully!\n\nComplete: {os.path.basename(complete_path)}\nIncomplete: {os.path.basename(incomplete_path)}"
            )
            root.destroy()
        except Exception as e:
            print(f"\n✗ Error: {str(e)}")
            root = tk.Tk()
            root.withdraw()
            messagebox.showerror("Error", str(e))
            root.destroy()


    # We combine it into the main table
    else:
        print(f"\n Please select the main table first")
        print(f"\n ..................") 
        file_path1 = select_file()

        if not file_path1:
            print("No file sleected")
            return
        
        print(f"\n Now select the corrected Housing Audit Workbook")
        print(f"\n ..................")
        file_path2 = select_file()

        if not file_path2:
            print("No file sleected")
            return
        
        # Get column mappings from user
        col_map = get_column_mapping()
        
        print(f"DEBUG: col_map = {col_map}")  # ADD THIS

        if not col_map:
            print("Operation cancelled")
            return
        
        try:
            # For security lets create and spit out a new file?
            new_file_path = combine_audit_to_main(file_path1, file_path2, col_map)
            print(f"\n Success!")
            print(f" - New file saved as : {new_file_path}")

            #show success dialog
            root = tk.Tk()
            root.withdraw()
            messagebox.showinfo(
                "Sucess",
                f"File updated successfully!\n\nNew file: {os.path.basename(new_file_path)}"
            )
            root.destroy()
        except Exception as e:
            print(f"\n✗ Error: {str(e)}")
            root = tk.Tk()
            root.withdraw()
            messagebox.showerror("Error", str(e))
            root.destroy()

    



if __name__ == "__main__":
    main()